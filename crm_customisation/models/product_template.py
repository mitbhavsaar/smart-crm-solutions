from odoo import models,fields,api
import json
import openpyxl


import logging
_logger = logging.getLogger(__name__)

class ProductTemplate(models.Model):
    _inherit = "product.template"

    raisin_type_id = fields.Many2one(
        'raisin.type',
        string="Raisin Type",
        required=True
    )
    is_raisin_category = fields.Boolean(
        string="Is Raisin Category",
        compute="_compute_is_raisin_category",
        store=True
    )
    @api.depends('categ_id.is_raisin')
    def _compute_is_raisin_category(self):
        for rec in self:
            rec.is_raisin_category = rec.categ_id.is_raisin
            
    @api.onchange('categ_id')
    def _onchange_categ_id(self):
        if self.categ_id and self.categ_id.is_raisin:
            if not self.raisin_type_id:
                self.raisin_type_id = self.categ_id.raisin_type_id
        elif not self._origin or self.raisin_type_id:
            self.raisin_type_id = False

class ProductCategory(models.Model):
    _inherit = "product.category"
    
    is_raisin = fields.Boolean(string="Is Raisin")
    raisin_type_id = fields.Many2one("raisin.type", string="Default Raisin Type")
    
    # Direct file upload field
    template_file = fields.Binary(
        string="Upload Template (Excel/CSV)",
        help="Upload your pre-made calculation sheet"
    )
    template_filename = fields.Char(string="Template Filename")
    
    # Spreadsheet template as a separate record
    template_spreadsheet_id = fields.Many2one(
        'crm.lead.spreadsheet',
        string="Template Spreadsheet",
        ondelete='set null'
    )
    
    has_template = fields.Boolean(
        string="Has Template", 
        compute='_compute_has_template'
    )
    
    def _compute_has_template(self):
        for category in self:
            # SAFER APPROACH: Use try-catch for template_spreadsheet_id access
            try:
                has_spreadsheet = bool(category.template_spreadsheet_id)
            except Exception:
                has_spreadsheet = False
                
            has_file = bool(category.template_file)
            category.has_template = has_spreadsheet or has_file
    
    def action_upload_and_create_template(self):
        """Upload Excel and convert to spreadsheet template"""
        self.ensure_one()
        
        if not self.template_file:
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'type': 'warning',
                    'message': 'Please upload a file first!',
                }
            }
        
        try:
            # Create new spreadsheet template
            new_spreadsheet = self.env['crm.lead.spreadsheet'].create({
                'name': f'{self.name} - Calculation Template',
                'category_id': self.id,
            })
            
            # Convert uploaded file to spreadsheet data
            spreadsheet_data = self._convert_file_to_spreadsheet(self.template_file, self.template_filename)
            if spreadsheet_data:
                new_spreadsheet.raw_spreadsheet_data = json.dumps(spreadsheet_data)
            
            # Link it to category
            self.template_spreadsheet_id = new_spreadsheet
            
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'type': 'success',
                    'message': f'Template uploaded successfully for {self.name}!',
                }
            }
        except Exception as e:
            _logger.error("Error creating template: %s", str(e))
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'type': 'error',
                    'message': f'Error creating template: {str(e)}',
                }
            }
    
    def _convert_file_to_spreadsheet(self, file_data, filename):
        """Convert Excel/CSV to Odoo spreadsheet format with Data Validation and Cell Shifting"""
        import base64
        from io import BytesIO
        import re
        
        # Import openpyxl and related modules
        try:
            import openpyxl
            from openpyxl.formula.translate import Translator
            from openpyxl.utils import get_column_letter, column_index_from_string
            from openpyxl.worksheet.datavalidation import DataValidation
        except ImportError:
            _logger.warning("openpyxl not installed, skipping Excel parsing")
            return None

        if not file_data or not filename:
            return None
            
        try:
            # Decode file
            file_content = base64.b64decode(file_data)
            
            # Basic spreadsheet structure
            spreadsheet_data = {
                'version': 1,
                'sheets': [{
                    'id': 'template_sheet',
                    'name': 'Template',
                    'colNumber': 26,
                    'rowNumber': 100,
                    'cells': {},
                    'merges': [],
                    'dataValidation': [], # Initialize data validation list
                }]
            }
            
            # Parse Excel file
            if filename.lower().endswith(('.xlsx', '.xls')):
                workbook = openpyxl.load_workbook(BytesIO(file_content), data_only=False) # data_only=False to get formulas
                sheet = workbook.active
                
                cells = {}
                validations = []
                
                # Configuration for shifting (Start at Column B, Row 6)
                # Target: B6 (Col 2, Row 6)
                # Source: A1 (Col 1, Row 1)
                # Shift: Col +1, Row +5
                SHIFT_COL = 1
                SHIFT_ROW = 5
                
                # Helper to shift cell reference (e.g., "A1" -> "B6")
                def shift_ref(ref):
                    try:
                        # Handle ranges like "A1:B2"
                        if ':' in ref:
                            start, end = ref.split(':')
                            return f"{shift_ref(start)}:{shift_ref(end)}"
                        
                        # Handle single cell "A1"
                        col_part = "".join(filter(str.isalpha, ref))
                        row_part = "".join(filter(str.isdigit, ref))
                        
                        if not col_part or not row_part:
                            return ref
                            
                        col_idx = column_index_from_string(col_part)
                        row_idx = int(row_part)
                        
                        new_col_idx = col_idx + SHIFT_COL
                        new_row_idx = row_idx + SHIFT_ROW
                        
                        return f"{get_column_letter(new_col_idx)}{new_row_idx}"
                    except Exception:
                        return ref

                # 1. Process Cells
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value is not None:
                            # Calculate new position
                            new_col_idx = cell.column + SHIFT_COL
                            new_row_idx = cell.row + SHIFT_ROW
                            new_col_letter = get_column_letter(new_col_idx)
                            new_cell_ref = f"{new_col_letter}{new_row_idx}"
                            
                            # Process content
                            cell_content = cell.value
                            
                            # Handle Formulas
                            if isinstance(cell_content, str) and cell_content.startswith('='):
                                try:
                                    # Translate formula
                                    original_ref = f"{get_column_letter(cell.column)}{cell.row}"
                                    cell_content = Translator(cell_content, origin=original_ref).translate_formula(new_cell_ref)
                                except Exception as e:
                                    _logger.warning(f"Failed to translate formula {cell_content}: {e}")
                            
                            # Store cell data
                            cells[new_cell_ref] = {
                                'content': str(cell_content),
                            }
                            
                            # Copy style/format if needed (simplified for now)
                            # if cell.number_format: ...

                # 2. Process Data Validations
                if hasattr(sheet, 'data_validations'):
                    for dv in sheet.data_validations.dataValidation:
                        # We only handle 'list' type for dropdowns for now
                        if dv.type == 'list':
                            # Parse ranges
                            sqref = dv.sqref
                            shifted_ranges = []
                            
                            # sqref can be a generic object or string in different openpyxl versions
                            # usually it's a MultiCellRange or string "A1:A5 B2:B2"
                            ranges_str = str(sqref).split()
                            
                            for rng in ranges_str:
                                shifted_ranges.append(shift_ref(rng))
                            
                            # Extract values
                            # formula1 might be '"Option1,Option2"' or 'Sheet2!$A$1:$A$5'
                            formula1 = dv.formula1
                            
                            rule = {
                                'type': 'list',
                                'values': []
                            }
                            
                            # Check if it's a direct list of values
                            if formula1.startswith('"') and formula1.endswith('"'):
                                # "Option1,Option2"
                                values_str = formula1.strip('"')
                                rule['values'] = values_str.split(',')
                            elif ',' in formula1 and not formula1.startswith('='):
                                # Option1,Option2 (sometimes without quotes)
                                rule['values'] = formula1.split(',')
                            else:
                                # It's a reference range (e.g. Sheet2!A1:A5)
                                # Odoo spreadsheet might not support cross-sheet references easily in this import format
                                # unless we resolve them to values or keep them as range references.
                                # For now, we'll try to keep it as a range source if Odoo supports it, 
                                # OR warn/skip. 
                                # Actually, Odoo spreadsheet JSON for validation usually expects explicit values 
                                # OR a range in the SAME sheet.
                                # If it's a reference, we might need to resolve it if it's static data.
                                # Let's try to store it as is for now, or maybe Odoo expects 'source' for range?
                                # Looking at Odoo source code (if I could), 'source' is used for range.
                                rule['source'] = formula1
                                # If we shifted the source cells (if they are on the same sheet), we should shift this too.
                                # But usually source lists are on a separate 'Data' sheet which we might not be shifting?
                                # If the source is on the SAME sheet, we must shift it.
                                if '!' not in formula1:
                                     rule['source'] = shift_ref(formula1)
                            
                            validations.append({
                                'ranges': shifted_ranges,
                                'rule': rule
                            })

                spreadsheet_data['sheets'][0]['cells'] = cells
                spreadsheet_data['sheets'][0]['dataValidation'] = validations
                
            elif filename.lower().endswith('.csv'):
                # CSV parsing (Keep existing logic but apply shift)
                try:
                    import csv
                    csv_content = file_content.decode('utf-8')
                    reader = csv.reader(csv_content.splitlines())
                    
                    cells = {}
                    # Shift for CSV too
                    SHIFT_COL = 1
                    SHIFT_ROW = 5
                    
                    for row_idx, row in enumerate(reader, start=1):
                        for col_idx, value in enumerate(row, start=1):
                            if value and value.strip():
                                # Shift
                                new_col_idx = col_idx + SHIFT_COL
                                new_row_idx = row_idx + SHIFT_ROW
                                
                                col_letter = self._get_column_letter(new_col_idx)
                                cell_ref = f"{col_letter}{new_row_idx}"
                                cells[cell_ref] = {'content': value.strip()}
                    
                    spreadsheet_data['sheets'][0]['cells'] = cells
                    
                except Exception as e:
                    _logger.error("Error parsing CSV: %s", str(e))
                    return None
            
            return spreadsheet_data
            
        except Exception as e:
            _logger.error("Error converting file to spreadsheet: %s", str(e))
            return None
    
    def _get_column_letter(self, col_idx):
        """Convert column index to letter (A, B, C, ... Z, AA, AB, etc.)"""
        letters = ''
        while col_idx > 0:
            col_idx, remainder = divmod(col_idx - 1, 26)
            letters = chr(65 + remainder) + letters
        return letters
    
    def action_open_template(self):
        """Open existing template spreadsheet"""
        self.ensure_one()
        if self.template_spreadsheet_id and self.template_spreadsheet_id.exists():
            return self.template_spreadsheet_id.action_open_spreadsheet()
        else:
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'type': 'warning',
                    'message': 'No template spreadsheet found for this category',
                }
            }
    
    def action_create_blank_template(self):
        """Create a blank template spreadsheet"""
        self.ensure_one()
        
        new_spreadsheet = self.env['crm.lead.spreadsheet'].create({
            'name': f'{self.name} - Blank Template',
            'category_id': self.id,
        })
        
        self.template_spreadsheet_id = new_spreadsheet
        
        return new_spreadsheet.action_open_spreadsheet()