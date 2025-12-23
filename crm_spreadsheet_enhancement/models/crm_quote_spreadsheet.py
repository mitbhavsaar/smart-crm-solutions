# -*- coding: utf-8 -*-
from odoo import api, fields, models, _
import json
import logging
import re
import uuid
import openpyxl
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.cell import range_boundaries
from openpyxl.formula.translate import Translator
import logging
_logger = logging.getLogger(__name__)


CRM_MATERIAL_LINE_BASE_FIELDS = [
    'product_template_id',  # Product
    'quantity',              # Quantity
    'product_uom_id',        # UOM
    'price',                 # Price
    'discount',              # Discount
    'tax_id',                # Tax
    'price_subtotal',        # Subtotal
]


class CrmLeadSpreadsheet(models.Model):
    _name = 'crm.lead.spreadsheet'
    _inherit = 'spreadsheet.mixin'
    _description = 'CRM Quotation Spreadsheet'

    name = fields.Char(required=True)
    lead_id = fields.Many2one('crm.lead', string="Opportunity", ondelete='cascade')
    sale_id = fields.Many2one('sale.order', string="Sale Order", ondelete='set null')
    company_id = fields.Many2one('res.company', default=lambda self: self.env.company)
    raw_spreadsheet_data = fields.Text("Raw Spreadsheet Data")

    # ------------------------------------------------------------------
    # ✅ CRITICAL: Override get_list_data (PUBLIC METHOD)
    # ------------------------------------------------------------------
    @api.model
    def get_list_data(self, model, list_id, field_names):
        """
        Override base spreadsheet method to handle dynamic attributes.
        """
        _logger.info(
            f"🟡 [Spreadsheet] get_list_data called: model={model}, "
            f"list_id={list_id}, fields={field_names}"
        )

        if model != 'crm.material.line':
            _logger.info(f"⚪ Not CRM model, using super: {model}")
            return super().get_list_data(model, list_id, field_names)

        try:
            line_id = int(list_id)
        except (ValueError, TypeError):
            _logger.error(f"❌ Invalid list_id: {list_id}")
            return []

        line = self.env['crm.material.line'].browse(line_id)
        if not line.exists():
            _logger.warning(f"❌ Material line {line_id} not found")
            return []

        _logger.info(f"✅ Found line {line_id}: {line.product_template_id.display_name}")

        # Get attributes_json FIRST
        attrs = line.attributes_json or {}
        _logger.info(f"📦 attributes_json: {attrs}")

        row = {"id": line.id}

        for field in field_names:
            if field in line._fields:
                # Standard Odoo field
                val = line[field]
                if hasattr(val, "display_name"):
                    row[field] = val.display_name
                else:
                    row[field] = val
                _logger.info(f"✅ Standard field '{field}' = '{row[field]}'")
            else:
                # Dynamic attribute from attributes_json
                row[field] = attrs.get(field, "")
                _logger.info(f"🔵 Dynamic field '{field}' = '{row[field]}'")

        _logger.info(f"🟡 Final row data: {row}")
        return [row]

    # ------------------------------------------------------------------
    # ✅ INTERNAL: _get_list_data (PRIVATE METHOD)
    # ------------------------------------------------------------------
    def _get_list_data(self, list_id):
        """
        Internal method for other operations
        """
        self.ensure_one()

        try:
            list_id_int = int(list_id)
        except (ValueError, TypeError):
            return []

        line = self.env['crm.material.line'].browse(list_id_int)
        if not line.exists():
            return []

        row = {
            'id': line.id,
            'product_template_id': line.product_template_id.display_name
            if line.product_template_id else '',
            'quantity': line.quantity or 0,
        }

        attrs = line.attributes_json or {}
        for key, value in attrs.items():
            row[key] = value

        return [row]

    # ------------------------------------------------------------------
    # OPEN FORMVIEW
    # ------------------------------------------------------------------
    def get_formview_action(self, access_uid=None):
        return self.action_open_spreadsheet()

    def action_open_spreadsheet(self):
        self.ensure_one()
        return {
            'type': 'ir.actions.client',
            'tag': 'action_crm_lead_spreadsheet',
            'params': {
                'spreadsheet_id': self.id,
                'model': 'crm.lead.spreadsheet',
            },
        }

    # ------------------------------------------------------------------
    # CREATE
    # ------------------------------------------------------------------
    @api.model_create_multi
    def create(self, vals_list):
        records = super().create(vals_list)
        for rec in records:
            if rec.lead_id and rec.lead_id.material_line_ids:
                for line in rec.lead_id.material_line_ids:
                    rec.with_context(
                        material_line_id=line.id
                    )._dispatch_insert_list_revision()
        return records

    # ------------------------------------------------------------------
    # JOIN SESSION
    # ------------------------------------------------------------------
    def join_spreadsheet_session(self, access_token=None):
        self.ensure_one()

        self._sync_sheets_with_material_lines()

        data = super().join_spreadsheet_session(access_token)
        data.update({
            'lead_id': self.lead_id.id if self.lead_id else False,
            'lead_display_name': self.lead_id.display_name if self.lead_id else False,
            'sheet_id': self.id,
        })

        spreadsheet_json = data.get('data') or {}
        lists = spreadsheet_json.get('lists') or {}
        sheets = spreadsheet_json.get('sheets') or []

        current_line_ids = set(self.lead_id.material_line_ids.ids) if self.lead_id else set()
        existing_list_ids = {int(list_id) for list_id in lists.keys() if list_id.isdigit()}

        missing_ids = current_line_ids - existing_list_ids
        removed_ids = existing_list_ids - current_line_ids

        # Add sheets
        for line_id in missing_ids:
            new_sheet = self._create_sheet_for_material_line(line_id)
            lists[str(line_id)] = new_sheet['list']
            sheets.append(new_sheet['sheet'])

        # Remove sheets
        if removed_ids:
            for rid in removed_ids:
                if str(rid) in lists:
                    del lists[str(rid)]
            sheets = [
                s for s in sheets
                if not any(str(rid) in json.dumps(s) for rid in removed_ids)
            ]

        spreadsheet_json['lists'] = lists
        spreadsheet_json['sheets'] = sheets

        # 🎯 REORDER SHEETS: Main sheets first, then auxiliary sheets
        # Main sheets have IDs like "sheet_123" (material line sheets)
        # Auxiliary sheets have IDs like "profile_master", "resin", "helper"
        main_sheets = []
        auxiliary_sheets = []
        
        for sheet in sheets:
            sheet_id = sheet.get('id', '')
            if sheet_id.startswith('sheet_'):
                # This is a main material line sheet
                main_sheets.append(sheet)
            else:
                # This is an auxiliary sheet (Profile, Resin, Helper, etc.)
                auxiliary_sheets.append(sheet)
        
        # Sort auxiliary sheets by priority
        def sheet_order(s):
            name = (s.get('name') or "").lower()
            if "profile master" in name or "profile" in name:
                return 1
            if "resin" in name:
                return 2
            if "helper" in name:
                return 3
            return 99
        
        auxiliary_sheets.sort(key=sheet_order)
        
        # Rebuild sheets array in correct order
        spreadsheet_json['sheets'] = main_sheets + auxiliary_sheets
        if main_sheets:
            first_main_sheet_id = main_sheets[0].get("id")
            data['active_sheet_id'] = first_main_sheet_id
            spreadsheet_json['active_sheet_id'] = first_main_sheet_id
            _logger.info(f"📌 Active sheet set to: {first_main_sheet_id}")
            _logger.info(f"📊 Sheet order: {[s.get('name') for s in spreadsheet_json['sheets']]}")

        # ✅ Preload data for ALL lists
        _logger.info("🔥 Preloading data for all lists...")
        for list_id, list_config in lists.items():
            try:
                line_id = int(list_id)
                line = self.env['crm.material.line'].browse(line_id)
                if line.exists():
                    columns = list_config.get('columns', [])
                    list_data = self.get_list_data('crm.material.line', list_id, columns)
                    _logger.info(f"✅ Preloaded list {list_id}: {list_data}")
            except Exception as e:
                _logger.error(f"❌ Failed to preload list {list_id}: {e}")

        data['data'] = spreadsheet_json
        self.raw_spreadsheet_data = json.dumps(spreadsheet_json)

        return data

    # ------------------------------------------------------------------
    # HELPER: Get Columns
    # ------------------------------------------------------------------
    def _get_material_line_columns(self, line):
        """
        Helper to construct columns for a material line sheet.
        Returns base fields + dynamic attributes (excluding duplicates).
        """
        # 1. Start with base fields
        columns = list(CRM_MATERIAL_LINE_BASE_FIELDS)
        
        # 2. Get dynamic attributes from attributes_json
        dynamic_keys = list(line.attributes_json.keys()) if isinstance(line.attributes_json, dict) else []
        
        # 3. Create a mapping of field names to their common variations
        # This helps filter out attributes that duplicate base fields
        field_variations = {
            'quantity': ['quantity', 'Quantity', 'qty', 'Qty'],
            'product_template_id': ['product', 'Product', 'product_template_id'],
            'product_uom_id': ['uom', 'UOM', 'uom_id', 'product_uom_id'],
            'price': ['price', 'Price'],
            'discount': ['discount', 'Discount'],
            'tax_id': ['tax', 'Tax', 'tax_id', 'taxes'],
            'price_subtotal': ['subtotal', 'Subtotal', 'price_subtotal'],
        }
        
        # Build a set of all variations to exclude
        exclude_variations = set()
        for field in columns:
            if field in field_variations:
                exclude_variations.update(field_variations[field])
        
        # Filter out dynamic keys that match any variation
        filtered_dynamic = [k for k in dynamic_keys if k not in exclude_variations]
        
        # 4. Priority ordering for product-specific attributes
        priority = []
        template = line.product_template_id
        if template:
            for ptal in template.attribute_line_ids:
                attr_name = ptal.attribute_id.name
                if attr_name in filtered_dynamic:
                    priority.append(attr_name)
                # Also check for UOM variants (e.g., "Width UOM")
                uom_name = f"{attr_name} UOM"
                if uom_name in filtered_dynamic:
                    priority.append(uom_name)
        
        # 5. Build ordered list of dynamic attributes
        ordered_dynamic = []
        remaining_dynamic = list(filtered_dynamic)
        
        # Add priority attributes first
        for p in priority:
            if p in remaining_dynamic:
                ordered_dynamic.append(p)
                remaining_dynamic.remove(p)
        
        # Add remaining attributes alphabetically
        ordered_dynamic.extend(sorted(remaining_dynamic))
        
        # 6. Combine base fields + ordered dynamic attributes
        columns.extend(ordered_dynamic)
        
        return columns

    # ------------------------------------------------------------------
    # EMPTY DATA (initial load)
    # ------------------------------------------------------------------
    def _empty_spreadsheet_data(self):
        data = super()._empty_spreadsheet_data() or {}
        data.setdefault('lists', {})
        data['sheets'] = []

        if not self.lead_id or not self.lead_id.material_line_ids:
            return data

        # 🔄 NEW: Collect all sheets first, then reorder
        main_sheets = []
        auxiliary_sheets = []
        
        for line in self.lead_id.material_line_ids:
            sheet_id = f"sheet_{line.id}"
            list_id = str(line.id)
            product_name = (line.product_template_id.display_name or "Item")[:31]

            columns = self._get_material_line_columns(line)

            # Check for template
            template_data = None
            if line.product_template_id.categ_id.spreadsheet_data:
                try:
                    template_data = json.loads(line.product_template_id.categ_id.spreadsheet_data)
                except Exception:
                    pass

            if template_data and template_data.get('sheets'):
                # 1. Main Sheet (Index 0)
                template_sheet = template_data['sheets'][0]
                # We need to deep copy to avoid modifying the cached template
                import copy
                sheet_json = copy.deepcopy(template_sheet)
                sheet_json['id'] = sheet_id
                sheet_json['name'] = product_name
                
                # ✅ SANITIZE: Ensure all cell content is string AND remove invalid format
                if 'cells' in sheet_json:
                    for cell_key, cell_val in sheet_json['cells'].items():
                        if 'content' in cell_val:
                            cell_val['content'] = str(cell_val['content']) if cell_val['content'] is not None else ""
                        if 'format' in cell_val:
                            del cell_val['format']
                            
                main_sheets.append(sheet_json)

                # 2. Auxiliary Sheets (Index 1+) - Collect for sorting
                for aux_sheet in template_data['sheets'][1:]:
                    # Check if sheet with this ID already exists
                    existing_ids = {s.get('id') for s in auxiliary_sheets}
                    if aux_sheet.get('id') not in existing_ids:
                        aux_copy = copy.deepcopy(aux_sheet)
                        # Sanitize
                        if 'cells' in aux_copy:
                            for cell_key, cell_val in aux_copy['cells'].items():
                                if 'content' in cell_val:
                                    cell_val['content'] = str(cell_val['content']) if cell_val['content'] is not None else ""
                                if 'format' in cell_val:
                                    del cell_val['format']
                        auxiliary_sheets.append(aux_copy)
            else:
                # Default behavior
                main_sheets.append({'id': sheet_id, 'name': product_name})

            data['lists'][list_id] = {
                'id': list_id,
                'model': 'crm.material.line',
                'columns': columns,
                'domain': [['id', '=', line.id]],
                'sheetId': sheet_id,
                'name': product_name,
                'context': {},
                'orderBy': [],
                'fieldMatching': {
                    'material_line_ids': {'chain': 'lead_id', 'type': 'many2one'},
                },
            }

        # 🎯 SORT auxiliary sheets: Profile Master → Resin → Helper
        def sheet_order(s):
            name = (s.get('name') or "").lower()
            if "profile master" in name or "profile" in name:
                return 1
            if "resin" in name:
                return 2
            if "helper" in name:
                return 3
            return 99

        auxiliary_sheets.sort(key=sheet_order)

        # ✅ Append in correct order: Main sheets first, then sorted auxiliaries
        data['sheets'].extend(main_sheets)
        data['sheets'].extend(auxiliary_sheets)

        return data
    
    def _shift_formula_rows(self, content, sheet_name, offset):
        if offset == 0 or not sheet_name:
            return content
        
        escaped_name = re.escape(sheet_name)
        # Match 'Sheet Name'!Ref or SheetName!Ref
        pattern = re.compile(f"(?:'({escaped_name})'|({escaped_name}))!(\\$?[A-Za-z]+)(\\$?)(\\d+)", re.IGNORECASE)
        
        def replace_func(match):
            prefix = match.group(1) or match.group(2)
            is_quoted = bool(match.group(1))
            col_part = match.group(3)
            row_anchor = match.group(4)
            row_num = int(match.group(5))
            new_row = row_num + offset
            sheet_part = f"'{prefix}'" if is_quoted else prefix
            return f"{sheet_part}!{col_part}{row_anchor}{new_row}"

        return pattern.sub(replace_func, content)

    def _fix_formula(self, content, original_name, new_name, all_sheet_names, main_sheet_info=None):
        if not content or not isinstance(content, str) or not content.startswith('='):
            return content
        
        # 0. Shift references to Main Sheet if needed
        if main_sheet_info:
            ms_name = main_sheet_info.get('name')
            ms_offset = main_sheet_info.get('offset', 0)
            if ms_name and ms_offset:
                content = self._shift_formula_rows(content, ms_name, ms_offset)

        # 1. Fix self-reference (rename current sheet in formulas)
        if original_name and new_name and original_name != new_name:
            # Replace 'Original Name'! with 'New Name'!
            content = content.replace(f"'{original_name}'!", f"'{new_name}'!")
            if ' ' not in original_name:
                content = content.replace(f"{original_name}!", f"'{new_name}'!")
        
        # 1.5 Fix references to main sheet in auxiliary sheets
        # When main "Costing" sheet is renamed to product name, update references
        if main_sheet_info and main_sheet_info.get('name') and main_sheet_info.get('new_name'):
            old_main = main_sheet_info['name']
            new_main = main_sheet_info['new_name']
            if old_main != new_main:
                # Update both quoted and unquoted references
                content = content.replace(f"'{old_main}'!", f"'{new_main}'!")
                if ' ' not in old_main:
                    # Use regex to avoid partial matches
                    import re
                    content = re.sub(
                        f"\\b{re.escape(old_main)}!",
                        f"'{new_main}'!",
                        content
                    )
        
        # 2. Ensure all sheet names in cross-sheet references are properly quoted
        # This prevents #NAME? errors
        if all_sheet_names:
            for sheet_name in all_sheet_names:
                if not sheet_name:
                    continue
                # If sheet name contains spaces or special chars, ensure it's quoted
                if ' ' in sheet_name or '-' in sheet_name or '.' in sheet_name:
                    # Replace unquoted sheet references with quoted ones
                    # Pattern: SheetName! but not 'SheetName'!
                    escaped_name = re.escape(sheet_name)
                    # Look for unquoted references
                    content = re.sub(
                        f"(?<!')[\\b]{escaped_name}(?=!)",
                        f"'{sheet_name}'",
                        content
                    )
        
        # 3. Ensure no double equals (prevents #BAD EXPR)
        if content.startswith("=="):
            content = content[1:]
        
        # 4. Fix common Excel → Odoo incompatibilities
        # Remove extra quotes that might cause issues
        content = content.replace("''", "'")
            
        return content
    def _get_validation_rule(self, val, all_sheet_names):
        """
        Convert Excel dropdown formula into correct Odoo Spreadsheet dataValidation rule.
        Supports:
            • "A,B,C" static list
            • =Sheet!$A$2:$A$1000 range dropdown
        """
        f1 = str(val.get('formula1', '')).strip()
        _logger.info(f"🔍 Processing Validation Rule: {f1}")

        # -------- STATIC LIST -------- ("A,B,C")
        if (f1.startswith('"') and f1.endswith('"')) or ("," in f1 and not f1.startswith("=")):
            raw = f1.strip('"')
            values = [v.strip() for v in raw.split(',')]
            _logger.info(f"   👉 Static List Detected = {values}")
            return {'type': 'range', 'values': values, 'style': 'arrow'}

        # -------- RANGE DROPDOWN -------- (=Sheet!$A$2:$A$100)
        if f1.startswith("="):
            f1 = f1[1:]  # remove "="

        clean = f1.replace("$", "").strip().strip('"')  

        # auto-quote ALWAYS (Odoo prefers quoted sheet names for validation ranges)
        if "!" in clean:
            sheet_part, range_part = clean.split("!", 1)
            # Strip existing quotes to avoid double quoting
            sheet_part = sheet_part.strip("'")
            clean = f"'{sheet_part}'!{range_part}"

        _logger.info(f"   👉 Final Clean Range for Odoo = {clean}")
        return {'type': 'range', 'range': clean, 'style': 'arrow'}


    def _get_sheet_populate_commands(
        self, sheet_json, sheet_id, row_offset=0, new_sheet_name=None,
        all_sheet_names=None, main_sheet_info=None
    ):
        """
        Populate sheet with merge, resize, formula-shift and Excel → Odoo dropdown conversion.
        """
        commands = []
        original_sheet_name_raw = sheet_json.get('name') or ""
        original_sheet_name = original_sheet_name_raw.lower()

        # 🔥 Detect reference sheets (No row offset for helper, resin, profile master etc.)
        reference_keywords = ["helper", "resin", "profile", "profile master", "master"]
        if any(k in original_sheet_name for k in reference_keywords):
            row_offset = 0
            col_offset = 0
        else:
            row_offset = 4
            col_offset = 0

        # -----------------------------------------
        # MERGES
        # -----------------------------------------
        for merge in sheet_json.get('merges', []):
            try:
                min_col, min_row, max_col, max_row = range_boundaries(merge)
                min_row += row_offset; max_row += row_offset
                min_col += col_offset; max_col += col_offset
                new_range = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
                commands.append({'type': 'ADD_MERGE', 'sheetId': sheet_id, 'target': [new_range]})
            except:
                pass

        # -----------------------------------------
        # COLUMN WIDTHS
        # -----------------------------------------
        for col_idx, col_data in sheet_json.get('cols', {}).items():
            commands.append({
                'type': 'RESIZE_COLUMNS_ROWS', 'sheetId': sheet_id,
                'dimension': 'COL', 'elements': [int(col_idx)],
                'size': col_data.get('width', 100) * 7
            })

        # -----------------------------------------
        # ROW HEIGHTS
        # -----------------------------------------
        for row_idx, row_data in sheet_json.get('rows', {}).items():
            commands.append({
                'type': 'RESIZE_COLUMNS_ROWS', 'sheetId': sheet_id,
                'dimension': 'ROW', 'elements': [int(row_idx) + row_offset],
                'size': row_data.get('size', 21)
            })

        # -----------------------------------------
        # 🔥 STEP 1: Build dropdown validation map (Excel → Odoo conversion)
        # -----------------------------------------
        cell_validations = {}
        if 'validations' in sheet_json:
            _logger.info(f"🔍 Found {len(sheet_json['validations'])} validation rules in template")
            for val in sheet_json['validations']:
                for rng in val.get('ranges', []):
                    try:
                        min_col, min_row, max_col, max_row = range_boundaries(rng)
                        for r in range(min_row, max_row + 1):
                            for c in range(min_col, max_col + 1):
                                cell_validations[(r - 1, c - 1)] = self._get_validation_rule(val, all_sheet_names)
                    except Exception as e:
                        _logger.warning(f"⚠️ Failed to parse validation range {rng}: {e}")

        _logger.info(f"📋 Total cells with validation: {len(cell_validations)}")

        # -----------------------------------------
        # 🔥 STEP 2: CELLS & FORMULAS (WITHOUT validation attached)
        # -----------------------------------------
        for cell_ref, cell_data in sheet_json.get('cells', {}).items():
            col_l = "".join(filter(str.isalpha, cell_ref))
            row_n = int("".join(filter(str.isdigit, cell_ref))) - 1
            col_n = column_index_from_string(col_l) - 1

            content = str(cell_data.get('content', '')) if cell_data.get('content') else ''

            # 🔁 Relocate formulas to new shifted position
            if content.startswith('='):
                try:
                    origin = f"{col_l}{row_n + 1}"
                    dest_col_letter = get_column_letter(col_n + col_offset + 1)
                    dest_ref = f"{dest_col_letter}{row_n + row_offset + 1}"
                    content = Translator(content, origin=origin).translate_formula(dest_ref)
                except:
                    pass
                content = self._fix_formula(content, original_sheet_name_raw, new_sheet_name, all_sheet_names, main_sheet_info)

            # ✅ Create cell command WITHOUT validation (clean separation)
            cmd = {
                'type': 'UPDATE_CELL',
                'sheetId': sheet_id,
                'col': col_n + col_offset, 
                'row': row_n + row_offset,
                'content': content,
            }
            commands.append(cmd)

        # -----------------------------------------
        # 🔥 STEP 3: Apply ALL validations as SEPARATE commands
        # -----------------------------------------
        _logger.info(f"🎯 Applying {len(cell_validations)} validations as separate commands...")
        
        validation_commands = []
        for (r, c), rule in cell_validations.items():
            target_col = c + col_offset
            target_row = r + row_offset
            cell_range = f"{get_column_letter(target_col + 1)}{target_row + 1}"
            
            # ✅ Build Odoo-compatible validation structure
            validation = None
            dv_id = f"dv_{sheet_id}_{target_row}_{target_col}"
            
            if rule.get('range'):
                # 🔥 Range-based dropdown (e.g., 'Profile Master'!A2:A1000)
                src = rule['range']
                
                # Clean range reference
                if src.startswith('"') and src.endswith('"'):
                    src = src[1:-1]
                if src.startswith("''") and not src.startswith("'''"):
                    src = src[1:]
                
                validation = {
                    'type': 'isValueInRange',
                    'values': [src],
                    'displayStyle': 'arrow'
                }
                _logger.info(f"✅ Range validation @ {cell_range}: {src}")
                
            elif rule.get('values'):
                # 🔥 Static list dropdown (e.g., ["A", "B", "C"])
                validation = {
                    'type': 'isValueInList',
                    'values': rule['values'],
                    'displayStyle': 'arrow'
                }
                _logger.info(f"✅ List validation @ {cell_range}: {rule['values']}")
            
            if validation:
                # ✅ Generate unique ID for this rule
               
                dv_id = f"dv_{uuid.uuid4().hex[:8]}"
                
                # ✅ CORRECT ODOO FORMAT (matching base implementation)
                validation_cmd = {
                    'type': 'ADD_DATA_VALIDATION_RULE',  # ← CORRECT command type
                    'sheetId': sheet_id,
                    'ranges': [cell_range],  # ← ranges at top level
                    'rule': {  # ← rule wrapper
                        'id': dv_id,  # ← id inside rule
                        'criterion': validation  # ← criterion inside rule
                    }
                }
                validation_commands.append(validation_cmd)
        
        # ✅ Append all validation commands AFTER cell content commands
        commands.extend(validation_commands)
        _logger.info(f"🎯 Added {len(validation_commands)} validation commands to dispatch queue")

        return commands




    # ------------------------------------------------------------------
    # INSERT REVISION (create new sheet on new line)
    # ------------------------------------------------------------------
    def _dispatch_insert_list_revision(self):
        self.ensure_one()

        line_id = self._context.get('material_line_id')
        if not line_id:
            return
        
        commands = []
        line = self.env['crm.material.line'].browse(line_id)
        if not line.exists():
            return

        sheet_id = f"sheet_{line.id}"
        list_id = str(line.id)
        product_name = (line.product_template_id.display_name or "Item")[:31]

        columns = self._get_material_line_columns(line)

        _logger.info(f"🔧 Creating sheet for line {line_id} with columns: {columns}")

        # Build column metadata with proper types
        columns_meta = []
        for col in columns:
            if col in self.env['crm.material.line']._fields:
                ftype = self.env['crm.material.line']._fields[col].type
            else:
                # Dynamic attribute - treat as char
                ftype = 'char'
            columns_meta.append({'name': col, 'type': ftype})

        # Get actual data now
        attrs = line.attributes_json or {}
        _logger.info(f"📦 Line {line_id} attributes_json: {attrs}")

        # Build the actual row data that will be inserted
        row_data = []
        for col_meta in columns_meta:
            field_name = col_meta['name']

            if field_name in line._fields:
                # Standard field
                val = line[field_name]
                if hasattr(val, 'display_name'):
                    cell_value = val.display_name
                else:
                    cell_value = val if val is not False else ''
            else:
                # Dynamic attribute
                cell_value = attrs.get(field_name, '')

            row_data.append(cell_value)
            _logger.info(f"  📝 {field_name} = {cell_value}")



        # Always create sheet and insert list first (Default behavior)
        # ✅ FORCE position 0 to ensure Main sheet is always first (before Aux sheets)
        commands.append({
            'type': 'CREATE_SHEET', 
            'sheetId': sheet_id, 
            'name': product_name,
            'position': 0 
        })
        
        commands.append({
            'type': 'REGISTER_ODOO_LIST',
            'listId': list_id,
            'model': 'crm.material.line',
            'columns': columns,
            'domain': [['id', '=', line.id]],
            'context': {},
            'orderBy': [],
        })
        
        commands.append({
            'type': 'RE_INSERT_ODOO_LIST',
            'sheetId': sheet_id,
            'col': 0,
            'row': 0,
            'id': list_id,
            'linesNumber': 1,
            'columns': columns_meta,
        })

        # Insert actual cell values immediately after creating the list
        for col_idx, (col_meta, cell_value) in enumerate(zip(columns_meta, row_data)):
            # 1. Data row
            commands.append({
                'type': 'UPDATE_CELL',
                'sheetId': sheet_id,
                'col': col_idx,
                'row': 1,  # Row 1 is data row (Row 0 is header)
                'content': str(cell_value)
                if cell_value not in (None, False, '') else '',
            })

            # 2. Header cleanup (for "__1" style names)
            field_name = col_meta['name']
            if "__" in field_name:
                display_name = field_name.split("__")[0]
                commands.append({
                    'type': 'UPDATE_CELL',
                    'sheetId': sheet_id,
                    'col': col_idx,
                    'row': 0,
                    'content': display_name,
                })

        # Add table formatting
        commands.append({
            'type': 'CREATE_TABLE',
            'sheetId': sheet_id,
            'tableType': 'static',
            'ranges': [{
                '_sheetId': sheet_id,
                '_zone': {
                    'top': 5,
                    'bottom': 500,
                    'left': 1,
                    'right': len(columns_meta),
                },
            }],
            'config': {
                'firstColumn': False,
                'hasFilters': True,
                'totalRow': False,
                'bandedRows': True,
                'styleId': 'TableStyleMedium5',
            },
        })

        # Check for template and append if exists
        template_data = None
        if line.product_template_id.categ_id.spreadsheet_data:
            try:
                template_data = json.loads(line.product_template_id.categ_id.spreadsheet_data)
            except Exception:
                pass

        if template_data and template_data.get('sheets'):
            _logger.info(f"📄 Merging template for line {line_id}")
            
            # Collect all sheet names for dynamic reference fixing
            all_sheet_names = [s.get('name') for s in template_data.get('sheets', [])]

            # 1. Main Sheet (Index 0) - Apply to sheet_id with offset
            main_sheet = template_data['sheets'][0]
            main_sheet_name = main_sheet.get('name')
            main_sheet_info = {
                'name': main_sheet_name, 
                'new_name': product_name,  # ✅ New name after merge
                'offset': 4
            }

            # Pass new_sheet_name=product_name to fix self-references
            template_cmds = self._get_sheet_populate_commands(
                main_sheet, 
                sheet_id, 
                row_offset=4, 
                new_sheet_name=product_name,
                all_sheet_names=all_sheet_names,
                main_sheet_info=main_sheet_info
            )
            commands.extend(template_cmds)

            # 2. Auxiliary Sheets (Index 1+)
            # We need to check if they exist in the spreadsheet already
            current_data = json.loads(self.raw_spreadsheet_data) if self.raw_spreadsheet_data else {}
            current_sheet_ids = {s.get('id') for s in current_data.get('sheets', [])}

            # 🎯 SORT auxiliary sheets: Profile Master → Resin → Helper → Others
            # This ensures correct tab order when spreadsheet opens
            aux_sheets = template_data['sheets'][1:]
            
            def sheet_order(s):
                name = (s.get('name') or "").lower()
                if "profile master" in name or "profile" in name:
                    return 1
                if "resin" in name:
                    return 2
                if "helper" in name:
                    return 3
                return 99

            aux_sheets.sort(key=sheet_order)

            for idx, aux_sheet in enumerate(aux_sheets, start=1):
                aux_id = aux_sheet.get('id')
                
                # Check if already exists in current data OR in pending revisions
                is_duplicate = False
                if aux_id in current_sheet_ids:
                    is_duplicate = True
                else:
                    # Check DB for pending commands creating this sheet
                    domain = [
                        ('res_id', '=', self.id),
                        ('res_model', '=', self._name),
                        ('commands', 'ilike', aux_id)
                    ]
                    if self.env['spreadsheet.revision'].search_count(domain):
                        is_duplicate = True

                if aux_id and not is_duplicate:
                    _logger.info(f"📄 Adding auxiliary sheet {aux_id} ({aux_sheet.get('name')})")
                    
                    # Create sheet with explicit position
                    commands.append({
                        'type': 'CREATE_SHEET', 
                        'sheetId': aux_id, 
                        'name': aux_sheet.get('name', 'Sheet')[:31],
                        'position': idx  # ✅ Force position 1, 2, 3...
                    })
                    
                    # Populate (No offset for reference sheets)
                    # Auxiliary sheets (Helper, Resin, Profile Master) are reference sheets
                    # They should NOT have row offset applied
                    aux_cmds = self._get_sheet_populate_commands(
                        aux_sheet, 
                        aux_id, 
                        row_offset=0,  # ✅ Changed from 4 to 0
                        new_sheet_name=aux_sheet.get('name'),
                        all_sheet_names=all_sheet_names,
                        main_sheet_info=main_sheet_info
                    )
                    commands.extend(aux_cmds)
                    
                    # Mark as added to avoid duplicates in this very loop (though unlikely)
                    current_sheet_ids.add(aux_id)

        # Final update command
        commands.append({'type': 'UPDATE_ODOO_LIST_DATA', 'listId': list_id})

        _logger.info(f"📤 Dispatching {len(commands)} commands for sheet {sheet_id}")
        self._dispatch_commands(commands)

    # ------------------------------------------------------------------
    # SYNC WITH MATERIAL LINES
    # ------------------------------------------------------------------
    def _sync_sheets_with_material_lines(self):
        self.ensure_one()
        if not self.lead_id:
            return

        data = json.loads(self.raw_spreadsheet_data) if self.raw_spreadsheet_data else {}
        current_sheets = data.get('sheets', [])
        current_lists = data.get('lists', {})
        current_line_ids = set(self.lead_id.material_line_ids.ids)

        # Remove deleted
        for sheet in current_sheets:
            sid = sheet.get('id')
            if sid and sid.startswith("sheet_"):
                try:
                    line_id = int(sid.replace("sheet_", ""))
                    if line_id not in current_line_ids:
                        self._delete_sheet_for_material_line(line_id)
                except Exception:
                    pass

        # Re-add missing OR update if columns changed
        existing_sheet_ids = {
            int(s['id'].replace('sheet_', ''))
            for s in current_sheets
            if s.get('id', '').startswith('sheet_')
        }

        for line in self.lead_id.material_line_ids:
            # 2. Check if sheet exists
            if line.id in existing_sheet_ids:
                list_id = str(line.id)
                if list_id in current_lists:
                    current_columns = current_lists[list_id].get('columns', [])

                    expected_columns = self._get_material_line_columns(line)

                    if current_columns != expected_columns:
                        _logger.info(
                            f"♻️ Columns changed for line {line.id}. "
                            f"Re-creating sheet."
                        )
                        self._delete_sheet_for_material_line(line.id)
                        self.with_context(
                            material_line_id=line.id
                        )._dispatch_insert_list_revision()
                        continue

            # 3. Create if missing
            if line.id not in existing_sheet_ids:
                self.with_context(
                    material_line_id=line.id
                )._dispatch_insert_list_revision()

    # ------------------------------------------------------------------
    # CREATE SHEET STRUCTURE
    # ------------------------------------------------------------------
    def _create_sheet_for_material_line(self, material_line_id):
        self.ensure_one()

        line = self.env['crm.material.line'].browse(material_line_id)
        if not line.exists():
            return {'sheet': {}, 'list': {}}

        sheet_id = f"sheet_{line.id}"
        list_id = str(line.id)
        name = (line.product_template_id.display_name or "Item")[:31]

        columns = self._get_material_line_columns(line)

        return {
            'sheet': {'id': sheet_id, 'name': name},
            'list': {
                'id': list_id,
                'model': 'crm.material.line',
                'columns': columns,
                'domain': [['id', '=', line.id]],
                'sheetId': sheet_id,
                'name': name,
                'context': {},
                'orderBy': [],
                'fieldMatching': {
                    'material_line_ids': {'chain': 'lead_id', 'type': 'many2one'},
                },
            },
        }

    # ------------------------------------------------------------------
    # DELETE SHEET
    # ------------------------------------------------------------------
    def _delete_sheet_for_material_line(self, material_line_id):
        sheet_id = f"sheet_{material_line_id}"
        list_id = str(material_line_id)

        commands = [
            {'type': 'DELETE_SHEET', 'sheetId': sheet_id},
            {'type': 'UNREGISTER_ODOO_LIST', 'listId': list_id},
        ]

        try:
            self._dispatch_commands(commands)
        except Exception:
            self._cleanup_deleted_sheets_from_data(material_line_id)

    def _cleanup_deleted_sheets_from_data(self, material_line_id):
        if not self.raw_spreadsheet_data:
            return
        try:
            data = json.loads(self.raw_spreadsheet_data)
            sid = f"sheet_{material_line_id}"
            if 'sheets' in data:
                data['sheets'] = [
                    s for s in data['sheets'] if s.get('id') != sid
                ]
            if 'lists' in data and str(material_line_id) in data['lists']:
                del data['lists'][str(material_line_id)]
            self.raw_spreadsheet_data = json.dumps(data)
        except Exception:
            pass

    # ------------------------------------------------------------------
    # MANUAL SYNC BUTTON
    # ------------------------------------------------------------------
    def action_sync_sheets(self):
        for spreadsheet in self:
            spreadsheet._sync_sheets_with_material_lines()
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'type': 'success',
                'message': _('Sheets synced with CRM Material Lines'),
                'next': {'type': 'ir.actions.act_window_close'},
            },
        }

    # ------------------------------------------------------------------
    # SPREADSHEET SELECTOR
    # ------------------------------------------------------------------
    @api.model
    def _get_spreadsheet_selector(self):
        return {
            'model': self._name,
            'display_name': _("CRM Quote Spreadsheets"),
            'sequence': 20,
            'allow_create': False,
        }

    # ------------------------------------------------------------------
    # DATA PROVIDER — used by spreadsheets
    # ------------------------------------------------------------------
    @api.model
    def get_crm_material_lines(self):
        self.ensure_one()
        if not self.lead_id:
            return []

        rows = []
        for line in self.lead_id.material_line_ids:
            row = {
                'id': line.id,
                'name': line.product_template_id.display_name
                if line.product_template_id else '',
                'quantity': line.quantity,
            }

            # Add dynamic JSON attributes
            attrs = line.attributes_json or {}
            for k, v in attrs.items():
                row[k] = v

            rows.append(row)

        return rows

    def getMainCrmMaterialLineLists(self):
        self.ensure_one()
        if not self.lead_id or not self.lead_id.material_line_ids:
            return []

        lists = []
        for line in self.lead_id.material_line_ids:
            columns = self._get_material_line_columns(line)

            lists.append({
                'id': str(line.id),
                'model': 'crm.material.line',
                'field_names': columns,
                'columns': columns,
                'name': line.product_template_id.display_name or f"Item {line.id}",
                'sheetId': f"sheet_{line.id}",
            })

        return lists