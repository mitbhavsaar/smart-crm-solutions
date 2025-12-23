# -*- coding: utf-8 -*-

from odoo import api, fields, models, _
from odoo.exceptions import UserError
import json
import logging

_logger = logging.getLogger(__name__)

SALES_ORDER_LINE_FIELDS = [
    'product_id',
    'product_uom_qty', 
    'price_unit',
    'width',
    'height',
    'length',
    'thickness',
]

class SaleOrderSpreadsheet(models.Model):
    _name = 'sale.order.spreadsheet'
    _inherit = 'spreadsheet.mixin'
    _description = 'Sales Order Spreadsheet'

    name = fields.Char(required=True)
    order_id = fields.Many2one('sale.order', ondelete='set null')
    company_id = fields.Many2one('res.company', default=lambda self: self.env.company)
    raw_spreadsheet_data = fields.Text("Raw Spreadsheet Data")

    # ✅ CRITICAL: Override get_list_data for Sales
    @api.model
    def get_list_data(self, model, list_id, field_names):
        """Get data for sale.order.line lists"""
        _logger.info(f"🟢 [Sales] get_list_data: model={model}, list_id={list_id}, fields={field_names}")
        
        if model != 'sale.order.line':
            return super().get_list_data(model, list_id, field_names)
        
        try:
            # Extract line ID
            if list_id.startswith('sales_'):
                line_id = int(list_id.replace('sales_', ''))
            else:
                line_id = int(list_id)
        except (ValueError, TypeError):
            _logger.error(f"❌ Invalid list_id: {list_id}")
            return []

        line = self.env['sale.order.line'].browse(line_id)
        if not line.exists():
            _logger.warning(f"❌ Sale order line {line_id} not found")
            return []

        _logger.info(f"✅ Found sale line {line_id}: {line.product_id.display_name}")
        
        row = {"id": line.id}

        for field in field_names:
            if field in line._fields:
                val = line[field]
                if hasattr(val, "display_name"):
                    row[field] = val.display_name
                else:
                    row[field] = val
                _logger.info(f"✅ Sale field '{field}' = '{row[field]}'")
            else:
                row[field] = ""

        return [row]

    def get_formview_action(self, access_uid=None):
        return self.action_open_spreadsheet()

    def action_open_spreadsheet(self):
        """Open sales spreadsheet"""
        self.ensure_one()
        return {
            'type': 'ir.actions.client',
            'tag': 'action_sale_order_spreadsheet', 
            'params': {
                'spreadsheet_id': self.id,
                'model': 'sale.order.spreadsheet', 
            },
        }

    # ✅ CRITICAL FIX: Convert CRM field syncs to Sales field syncs
    def _convert_crm_sheet_to_sales(self):
        """Convert CRM field syncs and lists to Sales format"""
        if not self.raw_spreadsheet_data:
            return

        try:
            data = json.loads(self.raw_spreadsheet_data)
        except Exception:
            return

        lists = data.get('lists', {}) or {}
        sheets = data.get('sheets', []) or []

        # ✅ FIELD MAPPING: CRM -> Sales
        FIELD_MAP = {
            'product_template_id': 'product_id',
            'quantity': 'product_uom_qty',
            'price': 'price_unit',
        }

        new_lists = {}
        new_sheets = []

        # Convert lists
        for list_key, lst in lists.items():
            if not isinstance(lst, dict):
                continue

            # Get record ID from domain
            crm_id = None
            domain = lst.get('domain', [])
            for condition in domain:
                if isinstance(condition, list) and len(condition) >= 3:
                    if condition[0] == 'id' and condition[1] == '=':
                        crm_id = int(condition[2])
                        break

            if list_key.startswith('crm_') and crm_id:
                new_key = f"sales_{crm_id}"
                new_lists[new_key] = {
                    'id': new_key,
                    'model': 'sale.order.line',  # ✅ CHANGED MODEL
                    'columns': SALES_ORDER_LINE_FIELDS,
                    'domain': [['id', '=', crm_id]],
                    'sheetId': f"sheet_sales_{crm_id}",
                    'name': lst.get('name', f"Item {crm_id}")[:31],
                    'context': {},
                    'orderBy': [],
                    'fieldMatching': {
                        'order_line': {'chain': 'order_id', 'type': 'many2one'},
                    },
                }
            elif list_key.startswith('sales_'):
                normalized = dict(lst)
                normalized['model'] = 'sale.order.line'  # ✅ ENSURE MODEL
                normalized['columns'] = SALES_ORDER_LINE_FIELDS
                new_lists[list_key] = normalized

        # ✅ CRITICAL: Convert sheets AND update fieldSyncs
        for sheet in sheets:
            if not isinstance(sheet, dict):
                continue
                
            sid = sheet.get('id', '')
            
            # Convert sheet IDs
            if sid.startswith('sheet_crm_'):
                try:
                    num = int(sid.replace('sheet_crm_', ''))
                except Exception:
                    continue
                
                new_sheet = {
                    'id': f"sheet_sales_{num}",
                    'name': sheet.get('name', f"Item {num}")[:31],
                    'cells': sheet.get('cells', {}) or {},
                    'figures': sheet.get('figures', []) or [],
                    'areGridLinesVisible': sheet.get('areGridLinesVisible', True),
                    'rowCount': sheet.get('rowCount', 1000),
                    'colCount': sheet.get('colCount', 26),
                }
                
                # ✅ CRITICAL: Update fieldSyncs in this sheet
                old_field_syncs = sheet.get('fieldSyncs', {}) or {}
                new_field_syncs = {}
                
                for cell_ref, field_sync in old_field_syncs.items():
                    new_field_sync = dict(field_sync)
                    
                    # Update list ID
                    if field_sync.get('listId', '').startswith('crm_'):
                        new_field_sync['listId'] = f"sales_{num}"
                    
                    # ✅ UPDATE FIELD NAME using mapping
                    old_field = field_sync.get('fieldName')
                    if old_field in FIELD_MAP:
                        new_field_sync['fieldName'] = FIELD_MAP[old_field]
                        _logger.info(f"🔄 Converted field: {old_field} -> {FIELD_MAP[old_field]}")
                    
                    new_field_syncs[cell_ref] = new_field_sync
                
                new_sheet['fieldSyncs'] = new_field_syncs
                new_sheets.append(new_sheet)
                
            elif sid.startswith('sheet_sales_'):
                new_sheets.append(sheet)

        data['lists'] = new_lists
        data['sheets'] = new_sheets

        try:
            self.raw_spreadsheet_data = json.dumps(data)
            _logger.info("✅ Successfully converted CRM spreadsheet to Sales format")
        except Exception as e:
            _logger.error(f"❌ Failed to save converted data: {e}")

    def _sync_order_lines_from_crm(self, crm_lead):
        """Sync order lines from CRM material lines"""
        try:
            order = self.order_id
            
            for material_line in crm_lead.material_line_ids:
                if material_line.product_id:
                    existing_line = order.order_line.filtered(
                        lambda l: l.product_id == material_line.product_id
                    )
                    
                    if not existing_line:
                        self.env['sale.order.line'].create({
                            'order_id': order.id,
                            'product_id': material_line.product_id.id,
                            'product_uom_qty': material_line.quantity or 1.0,
                            'price_unit': material_line.price or material_line.product_id.list_price,
                            'width': material_line.width or 0,
                            'height': material_line.height or 0,
                            'length': material_line.length or 0,
                            'thickness': material_line.thickness or 0,
                            'name': material_line.product_id.name,
                        })
                        
        except Exception as e:
            _logger.error(f"[ORDER_SYNC] Error: {str(e)}")

    def join_spreadsheet_session(self, access_token=None):
        """Join spreadsheet session - FIXED"""
        self.ensure_one()

        _logger.info(f"\n🟢 [SALES SESSION] Starting for {self.name}")

        # ✅ CRITICAL: Convert CRM data first
        converted_from_crm = False
        if self.raw_spreadsheet_data:
            crm_detected = any(marker in self.raw_spreadsheet_data 
                             for marker in ['"crm_', 'sheet_crm_', 'crm.material.line'])
            if crm_detected:
                _logger.info("🔄 Converting CRM data to Sales format...")
                self._convert_crm_sheet_to_sales()
                converted_from_crm = True
                _logger.info("✅ Conversion completed")

        # Sync sheets if needed
        should_sync = not self.raw_spreadsheet_data or converted_from_crm
        if should_sync:
            try:
                self._sync_sheets_with_order_lines()
            except Exception as e:
                _logger.error(f"❌ Sheet sync error: {e}")

        # Get base data
        data = super().join_spreadsheet_session(access_token)
        
        # Load spreadsheet data
        spreadsheet_json = {}
        if self.raw_spreadsheet_data:
            try:
                spreadsheet_json = json.loads(self.raw_spreadsheet_data)
                _logger.info(f"📊 Loaded: {len(spreadsheet_json.get('lists', {}))} lists")
            except Exception as e:
                _logger.error(f"❌ Data load error: {e}")
                spreadsheet_json = data.get('data') or {}
        else:
            spreadsheet_json = data.get('data') or {}

        # ✅ CRITICAL: Preload data for ALL sales lists
        lists = spreadsheet_json.get('lists', {})
        for list_id, list_config in lists.items():
            try:
                if list_id.startswith('sales_'):
                    line_id = int(list_id.replace('sales_', ''))
                else:
                    line_id = int(list_id)
                    
                line = self.env['sale.order.line'].browse(line_id)
                if line.exists():
                    columns = list_config.get('columns', [])
                    list_data = self.get_list_data('sale.order.line', list_id, columns)
                    _logger.info(f"✅ Preloaded sales list {list_id}")
            except Exception as e:
                _logger.error(f"❌ Failed to preload {list_id}: {e}")

        data['data'] = spreadsheet_json
        
        # Add sales context
        data.update({
            'order_id': self.order_id.id if self.order_id else False,
            'order_display_name': self.order_id.display_name if self.order_id else False,
            'sale_order_id': self.order_id.id if self.order_id else False,
            'sheet_id': self.id
        })

        _logger.info(f"🟢 [SALES SESSION] Completed\n")
        return data

    def _validate_list_domains(self, spreadsheet_data):
        """Validate list domains"""
        try:
            lists = spreadsheet_data.get('lists', {})
            
            for list_id, list_config in lists.items():
                domain = list_config.get('domain', [])
                record_id = None
                
                for condition in domain:
                    if (isinstance(condition, list) and len(condition) >= 3 and 
                        condition[0] == 'id' and condition[1] == '='):
                        record_id = condition[2]
                        break
                
                if record_id:
                    record = self.env['sale.order.line'].browse(record_id)
                    if not record.exists():
                        _logger.warning(f"⚠️ Invalid domain: List {list_id} -> Record {record_id}")
                        
        except Exception as e:
            _logger.error(f"❌ Validation error: {e}")

    def _dispatch_insert_list_revision(self):
        """Create and register sheet for sale order line"""
        self.ensure_one()
        line_id = self._context.get('order_line_id')
        if not line_id:
            return
        
        commands = []
        line = self.env['sale.order.line'].browse(line_id)
        if not line.exists():
            return

        sheet_id = f"sheet_sales_{line.id}"
        list_id = f"sales_{line.id}"
        product_name = (line.product_id.display_name or "Item")[:31]

        columns = [
            {'name': f, 'type': self.env['sale.order.line']._fields.get(f).type}
            for f in SALES_ORDER_LINE_FIELDS
        ]

        # Build row data
        row_data = []
        for col_meta in columns:
            field_name = col_meta['name']
            val = line[field_name]
            
            if hasattr(val, 'display_name'):
                cell_value = val.display_name
            else:
                cell_value = val if val is not False else ''
                
            row_data.append(cell_value)

        # Always create sheet and insert list first (Default behavior)
        commands.append({'type': 'CREATE_SHEET', 'sheetId': sheet_id, 'name': product_name})
        
        commands.append({
            'type': 'REGISTER_ODOO_LIST',
            'listId': list_id,
            'model': 'sale.order.line',
            'columns': SALES_ORDER_LINE_FIELDS,
            'domain': [['id', '=', line.id]],
            'context': {},
            'orderBy': [],
        })
        
        commands.append({
            'type': 'RE_INSERT_ODOO_LIST',
            'sheetId': sheet_id,
            'col': 0,
            'row': 0,
            'id': list_id,
            'linesNumber': 1,
            'columns': columns,
        })

        # Insert cell values
        for col_idx, (col_meta, cell_value) in enumerate(zip(columns, row_data)):
            commands.append({
                'type': 'UPDATE_CELL',
                'sheetId': sheet_id,
                'col': col_idx,
                'row': 1,
                'content': str(cell_value) if cell_value not in (None, False, '') else '',
            })

        commands.append({
            'type': 'CREATE_TABLE',
            'sheetId': sheet_id,
            'tableType': 'static',
            'ranges': [{
                '_sheetId': sheet_id,
                '_zone': {'top': 0, 'bottom': 1, 'left': 0, 'right': len(columns) - 1}
            }],
            'config': {
                'firstColumn': False,
                'hasFilters': True,
                'totalRow': False,
                'bandedRows': True,
                'styleId': 'TableStyleMedium5',
            }
        })

        # Check for template and append if exists
        template_data = None
        if line.product_id.product_tmpl_id.categ_id.spreadsheet_data:
            try:
                template_data = json.loads(line.product_id.product_tmpl_id.categ_id.spreadsheet_data)
            except Exception:
                pass

        if template_data:
            _logger.info(f"📄 Merging template for sale line {line_id} with offset")
            # Apply Template Content with OFFSET 4
            template_cmds = self._get_template_commands(sheet_id, template_data, row_offset=4)
            commands.extend(template_cmds)
        
        self._dispatch_commands(commands)

    @api.model_create_multi
    def create(self, vals_list):
        records = super().create(vals_list)
        for rec in records:
            if rec.order_id and rec.order_id.order_line and not rec.raw_spreadsheet_data:
                for line in rec.order_id.order_line:
                    rec.with_context(order_line_id=line.id)._dispatch_insert_list_revision()
        return records

    def _empty_spreadsheet_data(self):
        """Return sales spreadsheet structure"""
        data = super()._empty_spreadsheet_data() or {}
        data.setdefault('lists', {})
        data['sheets'] = []
        
        if not self.order_id or not self.order_id.order_line:
            return data

        for line in self.order_id.order_line:
            sheet_id = f"sheet_sales_{line.id}"
            list_id = f"sales_{line.id}"
            product_name = (line.product_id.display_name or "Untitled")[:31]

            # Check for template
            template_data = None
            if line.product_id.product_tmpl_id.categ_id.spreadsheet_data:
                try:
                    template_data = json.loads(line.product_id.product_tmpl_id.categ_id.spreadsheet_data)
                except Exception:
                    pass

            if template_data and template_data.get('sheets'):
                # Use template sheet
                template_sheet = template_data['sheets'][0]
                import copy
                sheet_json = copy.deepcopy(template_sheet)
                sheet_json['id'] = sheet_id
                sheet_json['name'] = product_name
                
                # ✅ SANITIZE: Ensure all cell content is string AND remove invalid format
                if 'cells' in sheet_json:
                    for cell_key, cell_val in sheet_json['cells'].items():
                        if 'content' in cell_val:
                            cell_val['content'] = str(cell_val['content']) if cell_val['content'] is not None else ""
                        if 'format' in cell_val:
                            del cell_val['format']

                data['sheets'].append(sheet_json)
            else:
                data['sheets'].append({
                    'id': sheet_id,
                    'name': product_name,
                })

            data['lists'][list_id] = {
                'id': list_id,
                'model': 'sale.order.line',
                'columns': SALES_ORDER_LINE_FIELDS,
                'domain': [['id', '=', line.id]],
                'sheetId': sheet_id,
                'name': product_name,
                'context': {},
                'orderBy': [],
                'fieldMatching': {
                    'order_line': {'chain': 'order_id', 'type': 'many2one'},
                },
            }
        return data

    def _get_template_commands(self, sheet_id, template_data, row_offset=0):
        """
        Generate commands to recreate the template sheet.
        """
        commands = []
        
        # 1. Cells
        if 'sheets' in template_data and template_data['sheets']:
            template_sheet = template_data['sheets'][0]
            
            # Merges
            from openpyxl.utils.cell import range_boundaries, get_column_letter
            for merge in template_sheet.get('merges', []):
                try:
                    min_col, min_row, max_col, max_row = range_boundaries(merge)
                    min_row += row_offset
                    max_row += row_offset
                    new_range = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
                    commands.append({
                        'type': 'ADD_MERGE',
                        'sheetId': sheet_id,
                        'target': [new_range]
                    })
                except Exception:
                    pass
                
            # Columns
            for col_idx, col_data in template_sheet.get('cols', {}).items():
                commands.append({
                    'type': 'RESIZE_COLUMNS_ROWS',
                    'sheetId': sheet_id,
                    'dimension': 'COL',
                    'elements': [int(col_idx)],
                    'size': col_data.get('width', 100) * 7 
                })
                
            # Rows
            for row_idx, row_data in template_sheet.get('rows', {}).items():
                commands.append({
                    'type': 'RESIZE_COLUMNS_ROWS',
                    'sheetId': sheet_id,
                    'dimension': 'ROW',
                    'elements': [int(row_idx) + row_offset],
                    'size': row_data.get('size', 21)
                })

            # Cells
            for cell_ref, cell_data in template_sheet.get('cells', {}).items():
                # Convert A1 to col/row
                col_letter = "".join(filter(str.isalpha, cell_ref))
                row_num = int("".join(filter(str.isdigit, cell_ref))) - 1
                from openpyxl.utils import column_index_from_string
                col_num = column_index_from_string(col_letter) - 1
                
                commands.append({
                    'type': 'UPDATE_CELL',
                    'sheetId': sheet_id,
                    'col': col_num,
                    'row': row_num + row_offset,
                    'content': str(cell_data.get('content', '')) if cell_data.get('content') is not None else '',
                    # 'format': cell_data.get('format'), 
                })
                
        return commands

    def _sync_sheets_with_order_lines(self):
        """Sync sheets with order lines"""
        self.ensure_one()
        
        if not self.order_id or self.raw_spreadsheet_data:
            return

        current_data = json.loads(self.raw_spreadsheet_data) if self.raw_spreadsheet_data else {}
        current_line_ids = set(self.order_id.order_line.ids)
        
        for line in self.order_id.order_line:
            self.with_context(order_line_id=line.id)._dispatch_insert_list_revision()
                
    def _create_sheet_for_order_line(self, order_line_id):
        """Create sheet for order line"""
        self.ensure_one()

        line = self.env['sale.order.line'].browse(order_line_id)
        if not line.exists():
            return None

        sheet_id = f"sheet_sales_{line.id}"
        list_id = f"sales_{line.id}"
        product_name = (line.product_id.display_name or f"Sales Item {line.id}")[:31]

        sheet_data = {
            'id': sheet_id,
            'name': product_name,
            'cells': {}, 
            'figures': [],
            'areGridLinesVisible': True,
            'rowCount': 1000,
            'colCount': 26,
        }

        list_data = {
            'id': list_id,
            'model': 'sale.order.line',
            'columns': SALES_ORDER_LINE_FIELDS,
            'domain': [['id', '=', line.id]],
            'sheetId': sheet_id,
            'name': product_name,
            'context': {},
            'orderBy': [],
            'fieldMatching': {
                'order_line': {'chain': 'order_id', 'type': 'many2one'},
            },
        }

        return {'sheet': sheet_data, 'list': list_data}

    def write_spreadsheet_data(self, data_json):
        """Save spreadsheet data"""
        self.ensure_one()

        try:
            data = json.loads(data_json)
        except Exception:
            return True

        self.raw_spreadsheet_data = data_json
        _logger.info(f"✅ Saved: {len(data.get('lists', {}))} lists, {len(data.get('sheets', []))} sheets")
        return True

    @api.model
    def _get_spreadsheet_selector(self):
        return {
            'model': self._name,
            'display_name': _("Sales Order Spreadsheets"),
            'sequence': 30,
            'allow_create': False,
        }

    def getMainSalesOrderLineLists(self):
        """Return sales order line lists"""
        self.ensure_one()
        if not self.order_id or not self.order_id.order_line:
            return []

        return [
            {
                'id': f"sales_{line.id}",
                'model': 'sale.order.line',
                'field_names': SALES_ORDER_LINE_FIELDS,
                'columns': SALES_ORDER_LINE_FIELDS,
                'name': line.product_id.display_name or f"Sales Item {line.id}",
                'sheetId': f"sheet_sales_{line.id}",
            }
            for line in self.order_id.order_line
        ]