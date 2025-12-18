# -*- coding: utf-8 -*-
from odoo import models, fields, api
import json
import openpyxl
import base64
from io import BytesIO
from openpyxl.utils import get_column_letter, column_index_from_string

class ProductCategory(models.Model):
    _inherit = "product.category"
    
    template_file = fields.Binary(string="Upload Calculation Template")
    template_filename = fields.Char(string="Template Filename")
    
    spreadsheet_data = fields.Text(
        string="Spreadsheet Data",
        compute='_compute_spreadsheet_data',
        store=True
    )
    
    @api.depends('template_file')
    def _compute_spreadsheet_data(self):
        for category in self:
            print("\n=========== DEBUG: _compute_spreadsheet_data ===========")
            print("CATEGORY:", category.name)
            print("template_file present?:", bool(category.template_file))
            
            if category.template_file:
                excel_data = category._convert_excel_to_spreadsheet(category.template_file)
                
                if excel_data:
                    category.spreadsheet_data = json.dumps(excel_data)
                    # small checksum for debugging
                    total_cells = sum(len(s.get('cells', {})) for s in excel_data.get('sheets', []))
                    total_merges = sum(len(s.get('merges', [])) for s in excel_data.get('sheets', []))
                    print("DEBUG: spreadsheet_data saved successfully! sheets:", len(excel_data.get('sheets', [])),
                          "cells:", total_cells, "merges:", total_merges)
                else:
                    print("DEBUG: excel_data conversion failed!")
            else:
                print("DEBUG: template_file is empty!")
                category.spreadsheet_data = False
            
            print("========================================================\n")
    

    def _parse_merge_range(self, range_str):
        """
        Convert 'A1:B3' -> dict {'top':0,'left':0,'bottom':2,'right':1}
        (0-based indices)
        """
        try:
            parts = range_str.split(':')
            if len(parts) == 1:
                # single cell treated as no-merge
                col = ''.join([c for c in parts[0] if c.isalpha()])
                row = ''.join([c for c in parts[0] if c.isdigit()])
                left = column_index_from_string(col) - 1
                top = int(row) - 1
                return {'top': top, 'left': left, 'bottom': top, 'right': left}
            start, end = parts
            col1 = ''.join([c for c in start if c.isalpha()])
            row1 = ''.join([c for c in start if c.isdigit()])
            col2 = ''.join([c for c in end if c.isalpha()])
            row2 = ''.join([c for c in end if c.isdigit()])
            left = column_index_from_string(col1) - 1
            top = int(row1) - 1
            right = column_index_from_string(col2) - 1
            bottom = int(row2) - 1
            return {'top': top, 'left': left, 'bottom': bottom, 'right': right}
        except Exception as e:
            # fallback: return None so caller can ignore
            print("DEBUG: parse_merge_range failed for", range_str, "error:", e)
            return None


    def _convert_excel_to_spreadsheet(self, file_data):
        """
        Convert uploaded XLSX (binary base64) into Odoo Spreadsheet JSON structure.
        Returns dict or None on failure.
        """
        try:
            print("\n******** DEBUG: Converting XLSX → Odoo Spreadsheet (openpyxl) ********")

            # decode and load workbook
            file_content = base64.b64decode(file_data)
            wb = openpyxl.load_workbook(BytesIO(file_content), data_only=False)

            spreadsheet = {
                "version": 16,
                "sheets": [],
                "revisionId": 1,
                "settings": {},
                "lists": {},
                "formats": {},
                "styles": {},
                "borders": {},
            }

            for sheet in wb.worksheets:
                # determine dimensions (fallbacks)
                max_row = sheet.max_row or 1
                max_col = sheet.max_column or 1

                # build sheet_json
                sheet_json = {
                    # unique id to avoid collision with sheet_<line.id>: prefix template_
                    "id": ("template_" + (sheet.title or "Sheet")).replace(" ", "_")[:60],
                    "name": (sheet.title or "Sheet")[:31],
                    # colNumber/rowNumber : use counts (Odoo expects integer)
                    "colNumber": int(max_col),
                    "rowNumber": max(int(max_row), 1000) if "profile master" in (sheet.title or "").lower() else int(max_row),
                    "cells": {},
                    "merges": [],
                    "rows": {},  # numeric-string keys: "0","1"
                    "cols": {},  # numeric-string keys: "0","1"
                }

                # -------------------------
                # merges: convert to numeric boxes
                # -------------------------
                try:
                    for merged in getattr(sheet, "merged_cells").ranges:
                        rng = str(merged)  # like 'A1:B3'
                        parsed = self._parse_merge_range(rng)
                        if parsed:
                            sheet_json["merges"].append(parsed)
                except Exception as e:
                    # if no merges or failure, ignore but log
                    print("DEBUG: reading merges failed for sheet", sheet.title, "error:", e)

                # -------------------------
                # column widths -> Odoo expects numeric index keys as strings
                # openpyxl.column_dimensions keys are letters like 'A'
                # convert: letter -> index-1 -> string key
                # -------------------------
                try:
                    for col_letter, col_dim in sheet.column_dimensions.items():
                        width = getattr(col_dim, "width", None)
                        if width is not None:
                            try:
                                idx = column_index_from_string(col_letter) - 1
                                sheet_json["cols"][str(idx)] = {"width": float(width)}
                            except Exception:
                                continue
                except Exception as e:
                    print("DEBUG: reading column_dimensions failed for sheet", sheet.title, "error:", e)

                # -------------------------
                # row heights -> numeric-string keys
                # -------------------------
                try:
                    for r_idx, row_dim in sheet.row_dimensions.items():
                        height = getattr(row_dim, "height", None)
                        if height is not None:
                            try:
                                sheet_json["rows"][str(int(r_idx) - 1)] = {"size": float(height)}
                            except Exception:
                                continue
                except Exception as e:
                    print("DEBUG: reading row_dimensions failed for sheet", sheet.title, "error:", e)

                # -------------------------
                # cells: iterate full rectangle so positions align
                # -------------------------
                for r in range(1, max_row + 1):
                    for c in range(1, max_col + 1):
                        cell = sheet.cell(row=r, column=c)
                        if cell is None:
                            continue
                        if cell.value is None:
                            # skip fully empty cells
                            continue

                        # key as A1 etc. Odoo expects A1-style keys inside cells dict
                        col_letter = get_column_letter(c)
                        key = f"{col_letter}{r}"
                        
                        # Handle ArrayFormula objects (can appear in any cell type)
                        from openpyxl.worksheet.formula import ArrayFormula
                        if isinstance(cell.value, ArrayFormula):
                            # Extract formula text from ArrayFormula object
                            raw = cell.value.text if hasattr(cell.value, 'text') else ""
                            if raw:
                                content = raw if raw.startswith('=') else '=' + raw
                            else:
                                # Empty array formula - skip
                                continue
                        # decide content: formula vs value
                        elif cell.data_type == 'f':
                            # Regular formula
                            raw = str(cell.value) if cell.value is not None else ""
                            # Skip formulas with __xludf (Excel dynamic array placeholders)
                            if '__xludf' in raw or '__xlud' in raw:
                                continue
                            content = raw if raw.startswith('=') else '=' + raw
                        else:
                            # preserve native python types for numbers/bool
                            v = cell.value
                            # Skip error values
                            if isinstance(v, str) and v.startswith('#'):
                                continue
                            # openpyxl may return datetime objects for dates — keep them as isoformat strings
                            try:
                                import datetime
                                if isinstance(v, (datetime.date, datetime.datetime)):
                                    content = v.isoformat()
                                else:
                                    content = v
                            except Exception:
                                content = str(v)
                        sheet_json["cells"][key] = {
                            "content": str(content) if content is not None else "",
                        }

                # -------------------------
                # Data Validations (Dropdowns)
                # -------------------------
                try:
                    if hasattr(sheet, 'data_validations') and sheet.data_validations:
                        sheet_json['validations'] = []
                        for dv in sheet.data_validations.dataValidation:
                            # We only care about LIST type for dropdowns usually
                            if dv.type == 'list':
                                # dv.sqref is a generic 'A1:A10 B1:B10' string or MultiCellRange
                                # We'll store it as a list of ranges
                                ranges = str(dv.sqref).split()
                                sheet_json['validations'].append({
                                    'type': 'list',
                                    'formula1': dv.formula1,
                                    'ranges': ranges,
                                    'showErrorMessage': dv.showErrorMessage,
                                    'showInputMessage': dv.showInputMessage,
                                })
                                print(f"DEBUG: Found validation in {sheet.title} | Type: {dv.type} | Formula: {dv.formula1} | Ranges: {ranges}")
                            else:
                                print(f"DEBUG: Ignored validation type: {dv.type} in {sheet.title}")
                except Exception as e:
                    print("DEBUG: reading data_validations failed for sheet", sheet.title, "error:", e)

                # append sheet
                spreadsheet["sheets"].append(sheet_json)
                print("DEBUG: parsed sheet:", sheet.title, "cells:", len(sheet_json["cells"]),
                      "merges:", len(sheet_json["merges"]),
                      "cols_meta:", len(sheet_json["cols"]),
                      "rows_meta:", len(sheet_json["rows"]))

            print("DEBUG: XLSX parsed. Total sheets:", len(spreadsheet["sheets"]))
            return spreadsheet

        except Exception as e:
            print("DEBUG: Excel conversion failed:", e)
            return None
